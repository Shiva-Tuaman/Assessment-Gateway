import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, date, time, timedelta
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
import hashlib
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import tempfile
import os
from io import BytesIO
# Email configuration
EMAIL_CONFIG = {
    'smtp_server': st.secrets["email"]["smtp_server"],
    'smtp_port': st.secrets["email"]["smtp_port"],
    'from_email': st.secrets["email"]["from_email"],
    'password': st.secrets["email"]["password"],
    'to_email': st.secrets["email"]["to_email"]
}
def generate_assessment_pdf(user_data, scores, interpretations, overall_assessment, total_possible, user_type="employee"):
    """Generate PDF report for assessment results"""
    try:
        buffer = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        doc = SimpleDocTemplate(buffer.name, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Tuaman Engineering Limited", title_style))
        story.append(Paragraph("Behavioral Competency Assessment Report", title_style))
        story.append(Spacer(1, 20))
        
        # User Information
        if user_type == "employee":
            user_info = [
                ['Employee ID:', user_data.get('employee_id', 'N/A')],
                ['Name:', user_data.get('employee_name', 'N/A')],
                ['Department:', user_data.get('department', 'N/A')],
                ['Assessment Date:', user_data.get('submit_date', 'N/A')]
            ]
        else:  # candidate
            user_info = [
                ['Candidate Code:', user_data.get('candidate_code', 'N/A')],
                ['Name:', user_data.get('full_name', 'N/A')],
                ['Position Applied:', user_data.get('position_applied', 'N/A')],
                ['Assessment Date:', user_data.get('submit_date', 'N/A')]
            ]
        
        # Create user info table
        user_table = Table(user_info, colWidths=[2*inch, 4*inch])
        user_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(user_table)
        story.append(Spacer(1, 20))
        
        # Scores section
        story.append(Paragraph("Assessment Scores", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        score_data = [['Competency', 'Score', 'Max Score', 'Percentage']]
        for comp, score in scores.items():
            max_score = total_possible.get(comp, 36)
            percentage = f"{(score/max_score)*100:.1f}%"
            score_data.append([comp, str(score), str(max_score), percentage])
        
        # Add total row
        total_score = sum(scores.values())
        total_max = sum(total_possible.values())
        total_percentage = f"{(total_score/total_max)*100:.1f}%"
        score_data.append(['TOTAL', str(total_score), str(total_max), total_percentage])
        
        score_table = Table(score_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(score_table)
        story.append(Spacer(1, 20))
        
        # Overall assessment
        story.append(Paragraph("Overall Assessment", styles['Heading2']))
        story.append(Paragraph(overall_assessment, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        buffer.close()
        
        return buffer.name
        
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")
        return None

def send_email_with_attachment(subject, body, attachment_path, attachment_name, cc_emails=None):
    """Send email with attachment"""
    try:
        st.write("📧 Starting email process...")
        
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['from_email']
        msg['To'] = EMAIL_CONFIG['to_email']
        msg['Subject'] = subject
        
        recipients = [EMAIL_CONFIG['to_email']]
        if cc_emails:
            cc_emails = [email.strip() for email in cc_emails if email.strip()]
            if cc_emails:
                msg['Cc'] = ', '.join(cc_emails)
                recipients.extend(cc_emails)
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach file if exists
        if attachment_path and os.path.exists(attachment_path):
            st.write(f"📎 Attaching file: {attachment_name}")
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{attachment_name}"'
            )
            msg.attach(part)
        
        # Connect to server
        st.write("🔗 Connecting to Gmail SMTP...")
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        
        st.write("🔐 Starting TLS encryption...")
        server.starttls()
        
        st.write("👤 Logging in...")
        server.login(EMAIL_CONFIG['from_email'], EMAIL_CONFIG['password'])
        
        st.write("📤 Sending email...")
        server.sendmail(EMAIL_CONFIG['from_email'], recipients, msg.as_string())
        server.quit()
        
        st.success(f"✅ Email sent successfully to: {', '.join(recipients)}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        st.error(f"❌ Authentication failed: {str(e)}")
        st.error("Check your Gmail App Password!")
        return False
    except smtplib.SMTPException as e:
        st.error(f"❌ SMTP error: {str(e)}")
        return False
    except Exception as e:
        st.error(f"❌ Email error: {str(e)}")
        st.error(f"Error type: {type(e).__name__}")
        return False

def show_email_preview(email_type, user_data, attachment_path=None, attachment_name=None):
    """Show email preview with editable fields"""
    st.subheader("📧 Email Preview")
    
    # Default email content
    if email_type == "excel_report":
        default_subject = "Employee Assessment Records - Excel Report"
        default_body = """Dear Team,

Please find attached the complete employee assessment records in Excel format.

This report contains all assessment data with conditional formatting for easy analysis.

Best regards,
Assessment System"""
    else:
        default_subject = f"Assessment Report - {user_data.get('employee_name', user_data.get('full_name', 'Employee'))}"
        default_body = f"""Dear Team,

Please find attached the assessment report.

Best regards,
Assessment System"""
    
    # Create form with unique key based on email type and timestamp
    import time
    form_key = f"email_form_{email_type}_{int(time.time())}"
    
    from_email = st.text_input("From:", value=EMAIL_CONFIG['from_email'], disabled=True, key=f"from_{form_key}")
    to_email = st.text_input("To:", value=EMAIL_CONFIG['to_email'], key=f"to_{form_key}")
    cc_emails = st.text_input("CC:", placeholder="email1@domain.com, email2@domain.com", key=f"cc_{form_key}")
    subject = st.text_input("Subject:", value=default_subject, key=f"subject_{form_key}")
    body = st.text_area("Message Body:", value=default_body, height=200, key=f"body_{form_key}")
    
    if attachment_name:
        st.info(f"📎 Attachment: {attachment_name}")
    
    if st.button("Send Email Now", type="primary", key=f"send_{form_key}"):
        if subject and body:
            cc_list = None
            if cc_emails:
                cc_list = [email.strip() for email in cc_emails.split(',') if email.strip()]
            
            # Update EMAIL_CONFIG
            EMAIL_CONFIG['to_email'] = to_email
            
            # Call the email function
            with st.spinner("Sending email..."):
                success = send_email_with_attachment(subject, body, attachment_path, attachment_name, cc_list)
            
            if success:
                st.success("✅ Email sent successfully!")
                
            else:
                st.error("❌ Failed to send email!")
        else:
            st.error("❌ Please fill in subject and message body.")
# Database setup
def init_database():
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    # Existing assessments table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS assessments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT NOT NULL,
            employee_name TEXT NOT NULL,
            department TEXT,
            assessment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            submit_date DATE,
            submit_time TIME,
            language TEXT,
            window_id INTEGER,
            accountability_score INTEGER,
            teamwork_score INTEGER,
            result_orientation_score INTEGER,
            communication_score INTEGER,
            adaptability_score INTEGER,
            integrity_score INTEGER,
            conflict_resolution_score INTEGER,
            total_score INTEGER,
            responses TEXT,
            interpretation TEXT,
            FOREIGN KEY (window_id) REFERENCES assessment_windows (id)
        )
    ''')
    
    # Check if columns exist, if not add them
    cursor.execute("PRAGMA table_info(assessments)")
    columns = [column[1] for column in cursor.fetchall()]
    
    if 'window_id' not in columns:
        cursor.execute('ALTER TABLE assessments ADD COLUMN window_id INTEGER')
    if 'submit_date' not in columns:
        cursor.execute('ALTER TABLE assessments ADD COLUMN submit_date DATE')
    if 'submit_time' not in columns:
        cursor.execute('ALTER TABLE assessments ADD COLUMN submit_time TIME')
    
    # Existing users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT UNIQUE NOT NULL,
            employee_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            department TEXT NOT NULL,
            user_type TEXT DEFAULT 'employee',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Existing assessment windows table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS assessment_windows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            window_name TEXT NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            start_time TIME NOT NULL,
            end_time TIME NOT NULL,
            is_active BOOLEAN DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT NOT NULL
        )
    ''')
    
    # NEW: Candidates table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_code TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            position_applied TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            expires_at DATETIME NOT NULL,
            is_active BOOLEAN DEFAULT 1
        )
    ''')
    
    # NEW: Candidate assessments table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS candidate_assessments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_code TEXT NOT NULL,
            full_name TEXT NOT NULL,
            position_applied TEXT,
            assessment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            submit_date DATE,
            submit_time TIME,
            language TEXT,
            accountability_score INTEGER,
            teamwork_score INTEGER,
            result_orientation_score INTEGER,
            communication_score INTEGER,
            adaptability_score INTEGER,
            integrity_score INTEGER,
            conflict_resolution_score INTEGER,
            total_score INTEGER,
            responses TEXT,
            interpretation TEXT,
            FOREIGN KEY (candidate_code) REFERENCES candidates (candidate_code)
        )
    ''')
    
    # NEW: Candidate admin table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS candidate_admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id TEXT UNIQUE NOT NULL,
            admin_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Insert default admins
    admin_password = hashlib.sha256("admin123".encode()).hexdigest()
    cursor.execute('''
        INSERT OR IGNORE INTO users (employee_id, employee_name, password_hash, department, user_type)
        VALUES (?, ?, ?, ?, ?)
    ''', ("admin", "Administrator", admin_password, "Administration", "admin"))
    
    candidate_admin_password = hashlib.sha256("candidateadmin123".encode()).hexdigest()
    cursor.execute('''
        INSERT OR IGNORE INTO candidate_admins (admin_id, admin_name, password_hash)
        VALUES (?, ?, ?)
    ''', ("candidateadmin", "Candidate Administrator", candidate_admin_password))
    
    conn.commit()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_user(employee_id, password):
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(password)
    cursor.execute('''
        SELECT employee_id, employee_name, department, user_type 
        FROM users 
        WHERE employee_id = ? AND password_hash = ?
    ''', (employee_id, password_hash))
    
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return {
            'employee_id': result[0],
            'employee_name': result[1],
            'department': result[2],
            'user_type': result[3]
        }
    return None

def create_user(employee_id, employee_name, password, department):
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(password)
    try:
        cursor.execute('''
            INSERT INTO users (employee_id, employee_name, password_hash, department)
            VALUES (?, ?, ?, ?)
        ''', (employee_id, employee_name, password_hash, department))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def generate_candidate_code():
    """Generate next candidate code"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) FROM candidates')
    count = cursor.fetchone()[0]
    conn.close()
    
    return f"TELCAN{count + 1:05d}"

def create_candidate(full_name, position_applied, password):
    """Create new candidate with 2-day expiry"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    candidate_code = generate_candidate_code()
    password_hash = hash_password(password)
    expires_at = datetime.now() + timedelta(days=2)
    
    try:
        cursor.execute('''
            INSERT INTO candidates (candidate_code, full_name, position_applied, password_hash, expires_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (candidate_code, full_name, position_applied, password_hash, expires_at))
        conn.commit()
        conn.close()
        return candidate_code
    except sqlite3.IntegrityError:
        conn.close()
        return None

def verify_candidate(candidate_code, password):
    """Verify candidate login and check expiry"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(password)
    cursor.execute('''
        SELECT candidate_code, full_name, position_applied, expires_at, is_active
        FROM candidates 
        WHERE candidate_code = ? AND password_hash = ?
    ''', (candidate_code, password_hash))
    
    result = cursor.fetchone()
    conn.close()
    
    if result:
        expires_at = datetime.fromisoformat(result[3])
        if datetime.now() > expires_at or not result[4]:
            return None  # Expired or inactive
        
        return {
            'candidate_code': result[0],
            'full_name': result[1],
            'position_applied': result[2],
            'expires_at': result[3]
        }
    return None

def verify_candidate_admin(admin_id, password):
    """Verify candidate admin login"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(password)
    cursor.execute('''
        SELECT admin_id, admin_name
        FROM candidate_admins 
        WHERE admin_id = ? AND password_hash = ?
    ''', (admin_id, password_hash))
    
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return {
            'admin_id': result[0],
            'admin_name': result[1],
            'user_type': 'candidate_admin'
        }
    return None

def has_candidate_taken_assessment(candidate_code):
    """Check if candidate has already taken assessment"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT COUNT(*) FROM candidate_assessments 
        WHERE candidate_code = ?
    ''', (candidate_code,))
    
    count = cursor.fetchone()[0]
    conn.close()
    
    return count > 0

def get_active_assessment_window():
    """Get currently active assessment window"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    # Use IST timezone
    from datetime import timezone, timedelta
    ist = timezone(timedelta(hours=5, minutes=30))
    
    current_datetime = datetime.now(ist)
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    
    # Convert time to string for SQLite compatibility
    current_time_str = current_time.strftime('%H:%M:%S')
    
    cursor.execute('''
        SELECT * FROM assessment_windows 
        WHERE is_active = 1 
        AND start_date <= ? AND end_date >= ?
        ORDER BY created_at DESC
    ''', (current_date, current_date))
    
    results = cursor.fetchall()
    conn.close()
    
    # Check time constraints in Python for more reliable comparison
    for result in results:
        start_time_str = result[4]  # start_time column
        end_time_str = result[5]    # end_time column
        
        # Convert string times to time objects for comparison
        try:
            start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
            end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
            
            # Check if current time is within the window
            if start_time <= current_time <= end_time:
                return {
                    'id': result[0],
                    'window_name': result[1],
                    'start_date': result[2],
                    'end_date': result[3],
                    'start_time': result[4],
                    'end_time': result[5],
                    'is_active': result[6]
                }
        except ValueError:
            # Skip if time format is invalid
            continue
    
    return None
def has_taken_assessment_in_window(employee_id, window_id):
    """Check if employee has already taken assessment in this window"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT COUNT(*) FROM assessments 
        WHERE employee_id = ? AND window_id = ?
    ''', (employee_id, window_id))
    
    count = cursor.fetchone()[0]
    conn.close()
    
    return count > 0

def create_assessment_window(window_name, start_date, end_date, start_time, end_time, created_by):
    """Create new assessment window"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    try:
        # Convert time objects to strings for SQLite compatibility
        start_time_str = start_time.strftime('%H:%M:%S') if hasattr(start_time, 'strftime') else str(start_time)
        end_time_str = end_time.strftime('%H:%M:%S') if hasattr(end_time, 'strftime') else str(end_time)
        
        cursor.execute('''
            INSERT INTO assessment_windows (window_name, start_date, end_date, start_time, end_time, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (window_name, start_date, end_date, start_time_str, end_time_str, created_by))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        conn.close()
        return False

def toggle_assessment_window(window_id, is_active):
    """Toggle assessment window active status"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE assessment_windows 
        SET is_active = ? 
        WHERE id = ?
    ''', (is_active, window_id))
    
    conn.commit()
    conn.close()

# Question bank with bilingual support (keeping existing questions)
QUESTIONS = {
    "Accountability": {
        "en": [
            {"type": "likert", "question": "I take full responsibility for my work outcomes, even when things go wrong.", "marks": 5},
            {"type": "situational", "question": "You missed a project deadline due to unexpected issues. What do you do?", 
             "options": ["Blame external factors", "Take responsibility and create recovery plan", "Wait for supervisor guidance", "Ignore and move to next task"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "Choose what describes you better:", 
             "options": ["I prefer clear instructions", "I take initiative without being asked"], "marks": 4},
            {"type": "likert", "question": "I consistently deliver on my commitments and promises.", "marks": 5},
            {"type": "situational", "question": "You discover an error in your completed work that no one else has noticed. You:", 
             "options": ["Keep quiet and hope no one finds out", "Immediately report and fix the error", "Wait to see if someone else catches it", "Fix it quietly without telling anyone"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "When facing challenges, I:", 
             "options": ["Look for someone else to handle it", "Take ownership and find solutions"], "marks": 4},
            {"type": "likert", "question": "I admit my mistakes openly and learn from them.", "marks": 5}
        ],
        "hi": [
            {"type": "likert", "question": "मैं अपने काम के परिणामों की पूरी जिम्मेदारी लेता हूं, भले ही चीजें गलत हो जाएं।", "marks": 5},
            {"type": "situational", "question": "आप अप्रत्याशित समस्याओं के कारण प्रोजेक्ट की समय सीमा चूक गए। आप क्या करते हैं?", 
             "options": ["बाहरी कारकों को दोष देना", "जिम्मेदारी लेना और रिकवरी प्लान बनाना", "सुपरवाइजर के मार्गदर्शन का इंतजार करना", "अनदेखा करके अगले काम पर जाना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "चुनें कि आपका बेहतर वर्णन क्या करता है:", 
             "options": ["मैं स्पष्ट निर्देश पसंद करता हूं", "मैं बिना कहे पहल करता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं अपनी प्रतिबद्धताओं और वादों को लगातार पूरा करता हूं।", "marks": 5},
            {"type": "situational", "question": "आप अपने पूरे किए गए काम में एक त्रुटि की खोज करते हैं जिसे किसी और ने नहीं देखा है। आप:", 
             "options": ["चुप रहना और उम्मीद करना कि कोई पता न लगाए", "तुरंत रिपोर्ट करना और त्रुटि ठीक करना", "देखना कि कोई और इसे पकड़ता है या नहीं", "चुपचाप इसे ठीक करना बिना किसी को बताए"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "चुनौतियों का सामना करते समय, मैं:", 
             "options": ["इसे संभालने के लिए किसी और को ढूंढता हूं", "स्वामित्व लेता हूं और समाधान खोजता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं अपनी गलतियों को खुले तौर पर स्वीकार करता हूं और उनसे सीखता हूं।", "marks": 5}
        ]
    },
    "Team Collaboration": {
        "en": [
            {"type": "likert", "question": "I actively contribute to team discussions and decision-making.", "marks": 5},
            {"type": "situational", "question": "A team member is struggling with their tasks. You:", 
             "options": ["Focus on your own work", "Offer help and support", "Report to supervisor", "Wait for them to ask for help"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "In team projects, I prefer to:", 
             "options": ["Work independently", "Collaborate closely with others"], "marks": 4},
            {"type": "likert", "question": "I respect and value diverse perspectives from team members.", "marks": 5},
            {"type": "situational", "question": "Your team has conflicting opinions on a project approach. You:", 
             "options": ["Push for your own idea", "Facilitate discussion to find common ground", "Stay neutral and let others decide", "Go with the majority opinion"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "When team goals conflict with personal goals, I:", 
             "options": ["Prioritize personal goals", "Put team goals first"], "marks": 4},
            {"type": "likert", "question": "I share knowledge and resources freely with my teammates.", "marks": 5}
        ],
        "hi": [
            {"type": "likert", "question": "मैं टीम की चर्चाओं और निर्णय लेने में सक्रिय रूप से योगदान देता हूं।", "marks": 5},
            {"type": "situational", "question": "एक टीम सदस्य अपने कार्यों के साथ संघर्ष कर रहा है। आप:", 
             "options": ["अपने काम पर ध्यान देना", "मदद और सहायता की पेशकश करना", "सुपरवाइजर को रिपोर्ट करना", "उनके मदद मांगने का इंतजार करना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "टीम प्रोजेक्ट्स में, मैं पसंद करता हूं:", 
             "options": ["स्वतंत्र रूप से काम करना", "दूसरों के साथ मिलकर काम करना"], "marks": 4},
            {"type": "likert", "question": "मैं टीम के सदस्यों के विविध दृष्टिकोणों का सम्मान और मूल्यांकन करता हूं।", "marks": 5},
            {"type": "situational", "question": "आपकी टीम में प्रोजेक्ट दृष्टिकोण पर विरोधाभासी राय हैं। आप:", 
             "options": ["अपने विचार के लिए जोर देना", "साझा आधार खोजने के लिए चर्चा की सुविधा देना", "तटस्थ रहना और दूसरों को निर्णय लेने देना", "बहुमत की राय के साथ जाना"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "जब टीम के लक्ष्य व्यक्तिगत लक्ष्यों से टकराते हैं, तो मैं:", 
             "options": ["व्यक्तिगत लक्ष्यों को प्राथमिकता देता हूं", "टीम के लक्ष्यों को पहले रखता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं अपने टीम के साथियों के साथ ज्ञान और संसाधनों को स्वतंत्र रूप से साझा करता हूं।", "marks": 5}
        ]
    },
    "Result Orientation": {
        "en": [
            {"type": "likert", "question": "I consistently focus on achieving measurable outcomes.", "marks": 5},
            {"type": "situational", "question": "You're working on a project with tight deadlines. You:", 
             "options": ["Work at your normal pace", "Prioritize tasks and work efficiently", "Ask for deadline extension", "Focus on perfection over completion"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "I am more motivated by:", 
             "options": ["The process of working", "Achieving specific results"], "marks": 4},
            {"type": "likert", "question": "I set clear, measurable goals for myself and track progress.", "marks": 5},
            {"type": "situational", "question": "A project is 80% complete but facing quality issues. You:", 
             "options": ["Rush to finish on time", "Address quality issues even if it delays completion", "Submit as is and fix later", "Seek guidance from supervisor"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "When facing obstacles, I:", 
             "options": ["Find alternative approaches", "Wait for conditions to improve"], "marks": 4},
            {"type": "likert", "question": "I celebrate achievements and learn from setbacks.", "marks": 5}
        ],
        "hi": [
            {"type": "likert", "question": "मैं लगातार मापने योग्य परिणाम प्राप्त करने पर ध्यान देता हूं।", "marks": 5},
            {"type": "situational", "question": "आप तंग समय सीमा वाले प्रोजेक्ट पर काम कर रहे हैं। आप:", 
             "options": ["अपनी सामान्य गति से काम करना", "कार्यों को प्राथमिकता देना और कुशलता से काम करना", "समय सीमा बढ़ाने के लिए कहना", "पूर्णता पर ध्यान देना बजाय समापन के"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "मैं अधिक प्रेरित होता हूं:", 
             "options": ["काम करने की प्रक्रिया से", "विशिष्ट परिणाम प्राप्त करने से"], "marks": 4},
            {"type": "likert", "question": "मैं अपने लिए स्पष्ट, मापने योग्य लक्ष्य निर्धारित करता हूं और प्रगति को ट्रैक करता हूं।", "marks": 5},
            {"type": "situational", "question": "एक प्रोजेक्ट 80% पूरा है लेकिन गुणवत्ता की समस्याओं का सामना कर रहा है। आप:", 
             "options": ["समय पर पूरा करने के लिए जल्दबाजी करना", "गुणवत्ता की समस्याओं को संबोधित करना भले ही इससे देरी हो", "जैसा है वैसा जमा करना और बाद में ठीक करना", "सुपरवाइजर से मार्गदर्शन लेना"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "बाधाओं का सामना करते समय, मैं:", 
             "options": ["वैकल्पिक दृष्टिकोण खोजता हूं", "स्थितियों के सुधरने का इंतजार करता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं उपलब्धियों का जश्न मनाता हूं और असफलताओं से सीखता हूं।", "marks": 5}
        ]
    },
    "Communication Skills": {
        "en": [
            {"type": "likert", "question": "I express my ideas clearly and concisely.", "marks": 5},
            {"type": "situational", "question": "You need to explain a complex technical concept to non-technical colleagues. You:", 
             "options": ["Use technical jargon", "Simplify and use analogies", "Refer them to documentation", "Ask a technical expert to explain"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "In meetings, I prefer to:", 
             "options": ["Listen more than speak", "Actively participate in discussions"], "marks": 4},
            {"type": "likert", "question": "I am an active listener who seeks to understand others' perspectives.", "marks": 5},
            {"type": "situational", "question": "A colleague seems confused by your instructions. You:", 
             "options": ["Blame them for not listening", "Repeat the same instructions louder", "Rephrase and confirm understanding", "Put everything in writing"], 
             "correct": 2, "marks": 6},
            {"type": "likert", "question": "I adapt my communication style based on my audience.", "marks": 5},
            {"type": "forced_choice", "question": "When presenting complex information, I:", 
             "options": ["Use technical details", "Simplify with examples"], "marks": 4}
        ],
        "hi": [
            {"type": "likert", "question": "मैं अपने विचारों को स्पष्ट और संक्षिप्त रूप से व्यक्त करता हूं।", "marks": 5},
            {"type": "situational", "question": "आपको गैर-तकनीकी सहयोगियों को एक जटिल तकनीकी अवधारणा समझानी है। आप:", 
             "options": ["तकनीकी शब्दजाल का उपयोग करना", "सरल बनाना और उदाहरण का उपयोग करना", "उन्हें दस्तावेज़ीकरण का संदर्भ देना", "किसी तकनीकी विशेषज्ञ से समझाने को कहना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "बैठकों में, मैं पसंद करता हूं:", 
             "options": ["बोलने से ज्यादा सुनना", "चर्चाओं में सक्रिय रूप से भाग लेना"], "marks": 4},
            {"type": "likert", "question": "मैं एक सक्रिय श्रोता हूं जो दूसरों के दृष्टिकोण को समझने की कोशिश करता हूं।", "marks": 5},
            {"type": "situational", "question": "एक सहयोगी आपके निर्देशों से भ्रमित लग रहा है। आप:", 
             "options": ["उन्हें न सुनने के लिए दोष देना", "वही निर्देश जोर से दोहराना", "दोबारा कहना और समझ की पुष्टि करना", "सब कुछ लिखित में देना"], 
             "correct": 2, "marks": 6},
            {"type": "likert", "question": "मैं अपने दर्शकों के आधार पर अपनी संचार शैली को अनुकूलित करता हूं।", "marks": 5},
            {"type": "forced_choice", "question": "जटिल जानकारी प्रस्तुत करते समय, मैं:", 
             "options": ["तकनीकी विवरण का उपयोग करता हूं", "उदाहरणों के साथ सरल बनाता हूं"], "marks": 4}
        ]
    },
    "Adaptability": {
        "en": [
            {"type": "likert", "question": "I embrace change as an opportunity for growth.", "marks": 5},
            {"type": "situational", "question": "Your company implements new software that changes your workflow. You:", 
             "options": ["Resist and prefer old methods", "Learn quickly and help others adapt", "Wait for formal training", "Complain about unnecessary changes"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "When plans change suddenly, I:", 
             "options": ["Feel stressed and overwhelmed", "Adjust and find new solutions"], "marks": 4},
            {"type": "likert", "question": "I remain calm and focused during unexpected situations.", "marks": 5},
            {"type": "situational", "question": "Your role responsibilities expand significantly. You:", 
             "options": ["Feel overwhelmed and resist", "Embrace the challenge and adapt", "Ask for additional compensation first", "Delegate to others"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "I prefer:", 
             "options": ["Predictable routines", "Variety and new challenges"], "marks": 4},
            {"type": "likert", "question": "I learn new skills quickly when required.", "marks": 5}
        ],
        "hi": [
            {"type": "likert", "question": "मैं परिवर्तन को विकास के अवसर के रूप में अपनाता हूं।", "marks": 5},
            {"type": "situational", "question": "आपकी कंपनी नया सॉफ्टवेयर लागू करती है जो आपके कार्यप्रवाह को बदल देता है। आप:", 
             "options": ["विरोध करना और पुराने तरीकों को पसंद करना", "जल्दी सीखना और दूसरों को अनुकूलित होने में मदद करना", "औपचारिक प्रशिक्षण का इंतजार करना", "अनावश्यक बदलावों की शिकायत करना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "जब योजनाएं अचानक बदल जाती हैं, तो मैं:", 
             "options": ["तनावग्रस्त और अभिभूत महसूस करता हूं", "समायोजित करता हूं और नए समाधान खोजता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं अप्रत्याशित स्थितियों के दौरान शांत और केंद्रित रहता हूं।", "marks": 5},
            {"type": "situational", "question": "आपकी भूमिका की जिम्मेदारियां काफी बढ़ जाती हैं। आप:", 
             "options": ["अभिभूत महसूस करना और विरोध करना", "चुनौती को अपनाना और अनुकूलित होना", "पहले अतिरिक्त मुआवजे के लिए कहना", "दूसरों को सौंपना"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "मैं पसंद करता हूं:", 
             "options": ["अनुमानित दिनचर्या", "विविधता और नई चुनौतियां"], "marks": 4},
            {"type": "likert", "question": "जब आवश्यक हो तो मैं नए कौशल जल्दी सीखता हूं।", "marks": 5}
        ]
    },
    "Integrity": {
        "en": [
            {"type": "likert", "question": "I always act according to my moral principles, even under pressure.", "marks": 5},
            {"type": "situational", "question": "You discover a billing error that benefits your company. You:", 
             "options": ["Keep quiet to benefit company", "Report it immediately", "Wait to see if client notices", "Discuss with colleagues first"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "When faced with ethical dilemmas, I:", 
             "options": ["Consider what others would do", "Follow my moral compass"], "marks": 4},
            {"type": "likert", "question": "I am honest about my capabilities and limitations.", "marks": 5},
            {"type": "situational", "question": "Your supervisor asks you to bend company policies for a client. You:", 
             "options": ["Comply without question", "Explain policy concerns and suggest alternatives", "Refuse directly", "Ask other colleagues what they would do"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "I believe:", 
             "options": ["Rules can be flexible when needed", "Principles should be consistently applied"], "marks": 4},
            {"type": "likert", "question": "I treat all people with respect and fairness.", "marks": 5}
        ],
        "hi": [
            {"type": "likert", "question": "मैं हमेशा अपने नैतिक सिद्धांतों के अनुसार काम करता हूं, दबाव में भी।", "marks": 5},
            {"type": "situational", "question": "आप एक बिलिंग त्रुटि की खोज करते हैं जो आपकी कंपनी को फायदा पहुंचाती है। आप:", 
             "options": ["कंपनी को फायदा पहुंचाने के लिए चुप रहना", "तुरंत रिपोर्ट करना", "देखना कि क्लाइंट नोटिस करता है या नहीं", "पहले सहयोगियों के साथ चर्चा करना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "नैतिक दुविधाओं का सामना करते समय, मैं:", 
             "options": ["विचार करता हूं कि दूसरे क्या करेंगे", "अपने नैतिक कम्पास का पालन करता हूं"], "marks": 4},
            {"type": "likert", "question": "मैं अपनी क्षमताओं और सीमाओं के बारे में ईमानदार हूं।", "marks": 5},
            {"type": "situational", "question": "आपका सुपरवाइजर आपसे एक क्लाइंट के लिए कंपनी की नीतियों को मोड़ने के लिए कहता है। आप:", 
             "options": ["बिना सवाल के पालन करना", "नीति की चिंताओं को समझाना और विकल्प सुझाना", "सीधे मना करना", "अन्य सहयोगियों से पूछना कि वे क्या करेंगे"], 
             "correct": 1, "marks": 6},
            {"type": "forced_choice", "question": "मैं मानता हूं:", 
             "options": ["नियम आवश्यकता पड़ने पर लचीले हो सकते हैं", "सिद्धांतों को लगातार लागू किया जाना चाहिए"], "marks": 4},
            {"type": "likert", "question": "मैं सभी लोगों के साथ सम्मान और निष्पक्षता से व्यवहार करता हूं।", "marks": 5}
        ]
    },
    "Conflict Resolution": {
        "en": [
            {"type": "likert", "question": "I handle conflicts constructively and seek win-win solutions.", "marks": 5},
            {"type": "situational", "question": "Two team members are in heated disagreement affecting project progress. You:", 
             "options": ["Let them work it out themselves", "Mediate and help find common ground", "Report to supervisor immediately", "Take sides with the person you agree with"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "In conflicts, I focus more on:", 
             "options": ["Winning the argument", "Finding mutual solutions"], "marks": 4},
            {"type": "likert", "question": "I remain neutral and objective when mediating disputes.", "marks": 5},
            {"type": "situational", "question": "You strongly disagree with your supervisor's decision. You:", 
             "options": ["Comply without expressing concerns", "Request private meeting to discuss concerns", "Publicly voice disagreement", "Seek support from other colleagues"], 
             "correct": 1, "marks": 6},
            {"type": "likert", "question": "I help others find common ground during disagreements.", "marks": 5},
            {"type": "forced_choice", "question": "When emotions run high in conflicts, I:", 
             "options": ["Wait for emotions to cool down", "Address emotional aspects first"], "marks": 4}
        ],
        "hi": [
            {"type": "likert", "question": "मैं संघर्षों को रचनात्मक तरीके से संभालता हूं और जीत-जीत के समाधान खोजता हूं।", "marks": 5},
            {"type": "situational", "question": "दो टीम सदस्य तीव्र असहमति में हैं जो प्रोजेक्ट की प्रगति को प्रभावित कर रहा है। आप:", 
             "options": ["उन्हें इसे खुद सुलझाने देना", "मध्यस्थता करना और साझा आधार खोजने में मदद करना", "तुरंत सुपरवाइजर को रिपोर्ट करना", "जिससे आप सहमत हैं उसका पक्ष लेना"], 
             "correct": 1, "marks": 7},
            {"type": "forced_choice", "question": "संघर्षों में, मैं अधिक ध्यान देता हूं:", 
             "options": ["बहस जीतने पर", "पारस्परिक समाधान खोजने पर"], "marks": 4},
            {"type": "likert", "question": "मैं विवादों की मध्यस्थता करते समय तटस्थ और वस्तुनिष्ठ रहता हूं।", "marks": 5},
            {"type": "situational", "question": "आप अपने सुपरवाइजर के निर्णय से दृढ़ता से असहमत हैं। आप:", 
             "options": ["चिंताओं को व्यक्त किए बिना पालन करना", "चर्चा के लिए निजी मीटिंग का अनुरोध करना", "सार्वजनिक रूप से असहमति व्यक्त करना", "अन्य सहकर्मियों से समर्थन लेना"], 
             "correct": 1, "marks": 6},
            {"type": "likert", "question": "मैं असहमति के दौरान दूसरों को साझा आधार खोजने में मदद करता हूं।", "marks": 5},
            {"type": "forced_choice", "question": "जब संघर्षों में भावनाएं तेज हो जाती हैं, तो मैं:", 
             "options": ["भावनाओं के शांत होने का इंतजार करता हूं", "पहले भावनात्मक पहलुओं को संबोधित करता हूं"], "marks": 4}
        ]
    }
}

# Scoring and interpretation logic
def calculate_scores(responses, language):
    scores = {}
    total_possible = {}
    
    for competency in QUESTIONS.keys():
        score = 0
        max_score = 0
        
        for i, question in enumerate(QUESTIONS[competency][language]):
            response = responses.get(f"{competency}_{i}", 0)
            max_score += question["marks"]
            
            if question["type"] == "likert":
                score += response * question["marks"] / 5
            elif question["type"] == "situational":
                if response == question["correct"]:
                    score += question["marks"]
            elif question["type"] == "forced_choice":
                if response == 1:  # Second option is usually the better choice
                    score += question["marks"]
        
        scores[competency] = round(score, 1)
        total_possible[competency] = max_score
    
    return scores, total_possible

def get_interpretation(scores, total_possible):
    interpretations = {}
    overall_categories = []
    
    for competency, score in scores.items():
        percentage = (score / total_possible[competency]) * 100
        
        if percentage >= 80:
            level = "Excellent"
            desc = "Demonstrates exceptional competency with consistent high performance"
        elif percentage >= 65:
            level = "Good"
            desc = "Shows strong competency with room for minor improvements"
        elif percentage >= 50:
            level = "Average"
            desc = "Displays adequate competency but needs focused development"
        elif percentage >= 35:
            level = "Below Average"
            desc = "Shows limited competency requiring significant improvement"
        else:
            level = "Poor"
            desc = "Demonstrates weak competency needing immediate attention"
        
        interpretations[competency] = {
            "level": level,
            "percentage": round(percentage, 1),
            "description": desc
        }
        overall_categories.append(level)
    
    # Overall assessment
    excellent_count = overall_categories.count("Excellent")
    good_count = overall_categories.count("Good")
    
    if excellent_count >= 5:
        overall = "High Performer"
    elif excellent_count + good_count >= 5:
        overall = "Strong Performer"
    elif overall_categories.count("Poor") >= 3:
        overall = "Needs Development"
    else:
        overall = "Average Performer"
    
    return interpretations, overall

def show_initial_selection():
    """Show initial selection between Existing Employee and New Candidate"""
    # Header with logo
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        try:
            st.image("Logo-TEL.png", width=300)
        except:
            st.write("Logo not found - please add Logo-TEL.png to project folder")
    
    # Main header
    st.markdown("""
    <div class="main-header">
        <h1>Welcome to Tuaman Engineering Limited</h1>
        <h2>Psychometric Assessment Portal</h2>
        <p>Comprehensive Behavioral Competency Evaluation System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # About section
    st.markdown("""
    <div style="background-color: #f8f9fa; padding: 2rem; border-radius: 10px; margin: 2rem 0;">
        <h3 style="color: #1f4e79; text-align: center;">About Our Assessment</h3>
        <p style="text-align: center; font-size: 16px; line-height: 1.6;">
            Our psychometric assessment evaluates key behavioral competencies essential for professional success. 
            The assessment covers areas including accountability, team collaboration, result orientation, 
            communication skills, adaptability, integrity, and conflict resolution.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Features section
        # Features section
    st.markdown("""
    <div style="margin: 2rem 0;">
        <div style="display: flex; justify-content: space-around; align-items: stretch; gap: 1rem;">
            <div style="flex: 1; text-align: center; padding: 2rem 1rem; background-color: white; border-radius: 10px; border: 1px solid #e0e0e0; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="font-size: 2.5rem; margin-bottom: 1rem;">🎯</div>
                <h4 style="color: #1f4e79; margin-bottom: 1rem; font-size: 1.2rem;">Comprehensive</h4>
                <p style="color: #666; font-size: 0.9rem; line-height: 1.4; margin: 0;">7 key competency areas evaluated</p>
            </div>
            <div style="flex: 1; text-align: center; padding: 2rem 1rem; background-color: white; border-radius: 10px; border: 1px solid #e0e0e0; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="font-size: 2.5rem; margin-bottom: 1rem;">🌐</div>
                <h4 style="color: #1f4e79; margin-bottom: 1rem; font-size: 1.2rem;">Multilingual</h4>
                <p style="color: #666; font-size: 0.9rem; line-height: 1.4; margin: 0;">Available in English and Hindi</p>
            </div>
            <div style="flex: 1; text-align: center; padding: 2rem 1rem; background-color: white; border-radius: 10px; border: 1px solid #e0e0e0; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="font-size: 2.5rem; margin-bottom: 1rem;">📊</div>
                <h4 style="color: #1f4e79; margin-bottom: 1rem; font-size: 1.2rem;">Detailed Reports</h4>
                <p style="color: #666; font-size: 0.9rem; line-height: 1.4; margin: 0;">Comprehensive analysis and insights</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    # Selection section
    st.markdown("""
    <div style="text-align: center; margin: 3rem 0 2rem 0;">
        <h3 style="color: #1f4e79;">Please select your category to continue</h3>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Card content
        st.markdown("""
        <div style="background-color: white; padding: 1.5rem; border-radius: 10px; border: 2px solid #1f4e79; margin: 1rem; height: 320px;">
            <h4 style="text-align: center; color: #1f4e79; margin-bottom: 1rem;">🏢 Existing Employee</h4>
            <p style="text-align: center; margin-bottom: 1rem; font-size: 14px;"><b>For current employees of Tuaman Engineering Limited</b></p>
            <ul style="margin: 1rem 0; padding-left: 1.5rem; font-size: 14px;">
                <li>Access your assessment dashboard</li>
                <li>View previous assessment results</li>
                <li>Take new assessments when available</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        # Image section
        col1_1, col1_2, col1_3 = st.columns([1, 2, 1])
        with col1_2:
            try:
                st.image("staff.png", width=120)
            except:
                st.markdown("""
                <div style="text-align: center; padding: 20px; background-color: #f0f2f6; border-radius: 10px; margin: 10px 0;">
                    <p style="color: #666;">👥 Staff Icon</p>
                </div>
                """, unsafe_allow_html=True)
        
        # Button
        if st.button("🏢 Existing Employee", use_container_width=True, type="primary"):
            st.session_state.user_category = "employee"
            st.rerun()
    
    with col2:
        # Card content
        st.markdown("""
        <div style="background-color: white; padding: 1.5rem; border-radius: 10px; border: 2px solid #1f4e79; margin: 1rem; height: 320px;">
            <h4 style="text-align: center; color: #1f4e79; margin-bottom: 1rem;">👤 New Candidate</h4>
            <p style="text-align: center; margin-bottom: 1rem; font-size: 14px;"><b>For job applicants and<br> new candidates</b></p>
            <ul style="margin: 1rem 0; padding-left: 1.5rem; font-size: 14px;">
                <li>Register for assessment access</li>
                <li>Complete your evaluation</li>
                <li>2-day access validity</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        # Image section
        col2_1, col2_2, col2_3 = st.columns([1, 2, 1])
        with col2_2:
            try:
                st.image("cv.png", width=120)
            except:
                st.markdown("""
                <div style="text-align: center; padding: 20px; background-color: #f0f2f6; border-radius: 10px; margin: 10px 0;">
                    <p style="color: #666;">📄 CV Icon</p>
                </div>
                """, unsafe_allow_html=True)
        
        # Button
        if st.button("👤 New Candidate", use_container_width=True, type="secondary"):
            st.session_state.user_category = "candidate"
            st.rerun()
    
    # Contact information
    st.markdown("""
    <div style="background-color: #e9ecef; padding: 1.5rem; border-radius: 10px; margin: 3rem 0 1rem 0; text-align: center;">
        <h4 style="color: #1f4e79;">Need Help?</h4>
        <p>Contact your HR department or system administrator for assistance</p>
        <p><strong>Tuaman Engineering Limited</strong><br>
        Head Office, Kolkata, West Bengal</p>
    </div>
    """, unsafe_allow_html=True)

def show_login_page():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col3:
        try:
            st.image("Logo-TEL.png", width=300)
        except:
            st.write("Logo not found - please add Logo-TEL.png to project folder")
    
    st.markdown("""
    <div class="main-header">
        <h1>Welcome to, </h1><h1>Tuaman Engineering Limited</h1><h1> Psychometric Assessment</h1>
        <p>Comprehensive Behavioral Competency Evaluation System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Back button
    if st.button("← Back to Main Menu"):
        st.session_state.user_category = None
        st.rerun()
    
    # Initialize forgot password state
    if 'show_forgot_password' not in st.session_state:
        st.session_state.show_forgot_password = False
    
    # Initialize active tab state
    if 'current_tab' not in st.session_state:
        st.session_state.current_tab = "login"
    
    # Initialize signup success state
    if 'signup_success' not in st.session_state:
        st.session_state.signup_success = False
    
    # Show forgot password form if requested
    if st.session_state.show_forgot_password:
        show_forgot_password_form()
        return
        
    # Show signup success message and redirect to login
    if st.session_state.signup_success:
        st.success("✅ Account created successfully! Please login with your credentials.")
        st.session_state.current_tab = "login"
        st.session_state.signup_success = False
    
    # Registration note for first-time users
    st.markdown("""
    <div style="background-color: #e3f2fd; padding: 1rem; border-radius: 8px; border-left: 4px solid #2196f3; margin: 1rem 0;">
        <h4 style="color: #1565c0; margin: 0 0 0.5rem 0;">📝 First Time User?</h4>
        <p style="margin: 0; color: #1976d2;">
            If you're accessing the assessment for the first time, please use the <strong>"Sign Up"</strong> button to register your account before logging in.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Tab selection buttons
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("🔑 Login", use_container_width=True, type="primary" if st.session_state.current_tab == "login" else "secondary"):
            st.session_state.current_tab = "login"
            st.rerun()
    with col2:
        if st.button("📝 Sign Up", use_container_width=True, type="primary" if st.session_state.current_tab == "signup" else "secondary"):
            st.session_state.current_tab = "signup"
            st.rerun()
    with col3:
        if st.button("👨‍💼 Admin Login", use_container_width=True, type="primary" if st.session_state.current_tab == "admin" else "secondary"):
            st.session_state.current_tab = "admin"
            st.rerun()
    
    st.markdown("---")
    
    # Show content based on selected tab
    if st.session_state.current_tab == "login":
        st.subheader("🔑 Login to Your Account")
        with st.form("login_form"):
            employee_id = st.text_input("Employee ID")
            password = st.text_input("Password", type="password")
            
            # Put Login and Forgot Password buttons on same line
            col1, col2 = st.columns([3, 1])
            with col1:
                login_button = st.form_submit_button("Login", type="primary")
            with col2:
                forgot_button = st.form_submit_button("Forgot Password?")
            
            if login_button:
                if employee_id and password:
                    user = verify_user(employee_id, password)
                    if user and user['user_type'] != 'admin':
                        st.session_state.user = user
                        st.session_state.authenticated = True
                        st.session_state.user_type = 'employee'
                        st.success(f"Welcome back, {user['employee_name']}!")
                        st.rerun()
                    else:
                        st.error("Invalid credentials. Please try again.")
                else:
                    st.error("Please fill in all fields.")
            
            if forgot_button:
                st.session_state.show_forgot_password = True
                st.rerun()
    
    elif st.session_state.current_tab == "signup":
        st.subheader("📝 Create New Account")
        # Password requirements
        st.markdown("""
        <div style="background-color: #fff3e0; padding: 1rem; border-radius: 8px; border-left: 4px solid #ff9800; margin: 1rem 0;">
            <h4 style="color: #e65100; margin: 0 0 0.5rem 0;">🔐 Password Requirements</h4>
            <div style="margin: 0.5rem 0;">
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Include uppercase letter (A-Z)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Include lowercase letter (a-z)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Include number (0-9)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Include special symbol (@, #, $)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Minimum 8 characters</span>
                </div>
            </div>
            <div style="margin-top: 1rem; padding: 0.5rem; background-color: #f5f5f5; border-radius: 5px;">
                <strong style="color: #e65100;">Example:</strong> <code style="background-color: #fff; padding: 0.2rem 0.4rem; border-radius: 3px;">Password@123</code>
            </div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("signup_form"):
            new_employee_id = st.text_input("Employee ID", key="signup_id")
            new_employee_name = st.text_input("Full Name", key="signup_name")
            new_password = st.text_input("Password", type="password", key="signup_password")
            confirm_password = st.text_input("Confirm Password", type="password", key="signup_confirm")
            new_department = st.selectbox(
                "Department",
                ["Engineering", "Procurement", "Business Development","Finance & Accounts","Human Resources & Administration","IT","Legal","Plant & Machinery","Production","Projects", "Project Management", "Quality Control","Quality Assurance", "Safety", "SAP","Strategy"],
                key="signup_dept"
            )
            signup_button = st.form_submit_button("Sign Up", type="primary")
            
            if signup_button:
                if all([new_employee_id, new_employee_name, new_password, confirm_password, new_department]):
                    if new_password != confirm_password:
                        st.error("Passwords do not match.")
                    else:
                        is_valid, message = validate_password(new_password)
                        if not is_valid:
                            st.error(message)
                        else:
                            if create_user(new_employee_id, new_employee_name, new_password, new_department):
                                st.session_state.signup_success = True
                                st.rerun()
                            else:
                                st.error("Employee ID already exists. Please use a different ID or login.")
                else:
                    st.error("Please fill in all fields.")
    
    elif st.session_state.current_tab == "admin":
        st.subheader("👨‍💼 Admin Login")
        with st.form("admin_login_form"):
            admin_employee_id = st.text_input("User Name", key="admin_id")
            admin_password = st.text_input("Password", type="password", key="admin_password")
            admin_login_button = st.form_submit_button("Admin Login", type="primary")
            
            if admin_login_button:
                if admin_employee_id and admin_password:
                    user = verify_user(admin_employee_id, admin_password)
                    if user and user['user_type'] == 'admin':
                        st.session_state.user = user
                        st.session_state.authenticated = True
                        st.session_state.user_type = 'employee'
                        st.success(f"Welcome Admin, {user['employee_name']}!")
                        st.rerun()
                    else:
                        st.error("Invalid admin credentials or insufficient privileges.")
                else:
                    st.error("Please fill in all fields.")   

def show_forgot_password_form():
    """Show the forgot password form"""
    
    st.subheader("🔑 Reset Password")
    st.markdown("""
    <div style="background-color: #fff3e0; padding: 1rem; border-radius: 8px; border-left: 4px solid #ffb74d; margin: 1rem 0;">
        <h4 style="color: #e65100; margin: 0 0 0.5rem 0;">🔐 Password Requirements</h4>
        <p style="margin: 0; color: #333; font-size: 14px;">
            • Start with uppercase letter<br>
            • Contain lowercase letters<br>
            • Include special symbol (@, #, $)<br>
            • Minimum 8 characters<br>
            <strong>Example:</strong> Password@123
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.info("Enter your Employee ID to reset your password")
    
    # Back to login button
    if st.button("← Back to Login"):
        st.session_state.show_forgot_password = False
        st.rerun()
    
    with st.form("forgot_password_form"):
        reset_employee_id = st.text_input("Employee ID", key="reset_emp_id")
        new_reset_password = st.text_input("New Password", type="password", key="reset_new_password")
        confirm_reset_password = st.text_input("Confirm New Password", type="password", key="reset_confirm_password")
        reset_button = st.form_submit_button("Reset Password", type="primary")
        
        if reset_button:
            if all([reset_employee_id, new_reset_password, confirm_reset_password]):
                employee_name = verify_user_exists(reset_employee_id)
                if not employee_name:
                    st.error("Employee ID not found.")
                elif new_reset_password != confirm_reset_password:
                    st.error("Passwords do not match.")
                else:
                    is_valid, message = validate_password(new_reset_password)
                    if not is_valid:
                        st.error(message)
                    else:
                        if reset_user_password(reset_employee_id, new_reset_password):
                            st.success(f"Password reset successful for {employee_name}! You can now login with your new password.")
                            st.session_state.show_forgot_password = False
                            st.rerun()
                        else:
                            st.error("Failed to reset password. Please try again.")
            else:
                st.error("Please fill in all fields.")
    
    # Password requirements reminder
       

def reset_user_password(employee_id, new_password):
    """Reset employee password"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(new_password)
    try:
        cursor.execute('''
            UPDATE users 
            SET password_hash = ? 
            WHERE employee_id = ?
        ''', (password_hash, employee_id))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return True
        else:
            conn.close()
            return False
    except Exception:
        conn.close()
        return False

def reset_candidate_password(candidate_code, new_password):
    """Reset candidate password"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    password_hash = hash_password(new_password)
    try:
        cursor.execute('''
            UPDATE candidates 
            SET password_hash = ? 
            WHERE candidate_code = ? AND is_active = 1
        ''', (password_hash, candidate_code))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return True
        else:
            conn.close()
            return False
    except Exception:
        conn.close()
        return False

def verify_user_exists(employee_id):
    """Check if employee exists"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT employee_name FROM users WHERE employee_id = ?', (employee_id,))
    result = cursor.fetchone()
    conn.close()
    
    return result[0] if result else None

def verify_candidate_exists(candidate_code):
    """Check if candidate exists and is active"""
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT full_name FROM candidates 
        WHERE candidate_code = ? AND is_active = 1 AND expires_at > datetime('now')
    ''', (candidate_code,))
    result = cursor.fetchone()
    conn.close()
    
    return result[0] if result else None
def validate_password(password):
    """Validate password according to requirements"""
    import re
    
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    
    if not password[0].isupper():
        return False, "Password must start with an uppercase letter."
    
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least one lowercase letter."
    
    if not re.search(r'[@#$]', password):
        return False, "Password must contain at least one special symbol (@, #, or $)."
    
    return True, "Password is valid."
def show_footer():
    """Display copyright footer on every page"""
    st.markdown("""
    <div style="
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        background-color: #f8f9fa;
        text-align: center;
        padding: 10px 0;
        border-top: 1px solid #dee2e6;
        font-size: 12px;
        color: #6c757d;
        z-index: 999;
    ">
        <p style="margin: 0;">© 2025 Tuaman Engineering Limited</p>
        <p style="margin: 0;">Head Office, Kolkata, West Bengal</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Add bottom padding to main content to prevent overlap
    st.markdown("""
    <style>
    .main .block-container {
        padding-bottom: 80px;
    }
    </style>
    """, unsafe_allow_html=True)
def show_candidate_login_page():
    """Show candidate login/signup page"""
    col1, col2, col3 = st.columns([1, 2, 1])
    with col3:
        try:
            st.image("Logo-TEL.png", width=300)
        except:
            st.write("Logo not found - please add Logo-TEL.png to project folder")
    
    st.markdown("""
    <div class="main-header">
        <h1>New Candidate Portal</h1>
        <p>Assessment access valid for 2 days from registration</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Back button
    if st.button("← Back to Main Menu"):
        st.session_state.user_category = None
        st.rerun()
    
    # Registration note for first-time candidates
    st.markdown("""
    <div style="background-color: #e3f2fd; padding: 1rem; border-radius: 8px; border-left: 4px solid #2196f3; margin: 1rem 0;">
        <h4 style="color: #1565c0; margin: 0 0 0.5rem 0;">📝 First Time Candidate?</h4>
        <p style="margin: 0; color: #1976d2;">
            If you're applying for the first time, please use the <strong>"New Registration"</strong> tab to create your candidate account before logging in.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["Candidate Login", "New Registration", "Admin Login"])
    
    with tab1:
        st.subheader("Login to Your Account")
        with st.form("candidate_login_form"):
            candidate_code = st.text_input("Candidate Code (e.g., TELCAN00001)")
            password = st.text_input("Password", type="password")
            
            # Login button with forgot password button beside it
            col1, col2 = st.columns([2, 1.5])
            with col1:
                login_button = st.form_submit_button("Login", type="primary")
            
            
            if login_button:
                if candidate_code and password:
                    candidate = verify_candidate(candidate_code, password)
                    if candidate:
                        st.session_state.user = candidate
                        st.session_state.authenticated = True
                        st.session_state.user_type = 'candidate'
                        st.success(f"Welcome, {candidate['full_name']}!")
                        st.rerun()
                    else:
                        st.error("Invalid credentials or account expired.")
                else:
                    st.error("Please fill in all fields.")
            
    
    with tab2:
        st.subheader("New Candidate Registration")
        st.markdown("""
        <div style="background-color: #fff3e0; padding: 1.5rem; border-radius: 10px; border: 1px solid #ffb74d; margin: 1rem 0;">
            <h4 style="color: #e65100; margin: 0 0 1rem 0;">🔐 Password Requirements</h4>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 0.5rem;">
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Start with uppercase letter</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Contain lowercase letters</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Include special symbol (@, #, $)</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <span style="color: #4caf50; margin-right: 0.5rem;">✓</span>
                    <span style="color: #333;">Minimum 8 characters</span>
                </div>
            </div>
            <div style="margin-top: 1rem; padding: 0.5rem; background-color: #f5f5f5; border-radius: 5px;">
                <strong style="color: #e65100;">Example:</strong> <code style="background-color: #fff; padding: 0.2rem 0.4rem; border-radius: 3px;">Password@123</code>
            </div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("candidate_signup_form"):
            full_name = st.text_input("Full Name")
            position_applied = st.text_input("Position Applied For")
            new_password = st.text_input("Create Password", type="password")
            confirm_password = st.text_input("Confirm Password", type="password")
            signup_button = st.form_submit_button("Register", type="primary")
            
            if signup_button:
                if all([full_name, position_applied, new_password, confirm_password]):
                    if new_password != confirm_password:
                        st.error("Passwords do not match.")
                    else:
                        is_valid, message = validate_password(new_password)
                        if not is_valid:
                            st.error(message)
                        else:
                            candidate_code = create_candidate(full_name, position_applied, new_password)
                            if candidate_code:
                                st.success(f"Registration successful! Your Candidate Code is: **{candidate_code}**")
                                st.info("⚠️ Please save this code. Your access will expire in 2 days.")
                                st.info("You can now login using your candidate code and password.")
                            else:
                                st.error("Registration failed. Please try again.")
                else:
                    st.error("Please fill in all fields.")
    
    with tab3:
        st.subheader("Candidate Admin Login")
        with st.form("candidate_admin_login_form"):
            admin_id = st.text_input("Admin ID")
            admin_password = st.text_input("Admin Password", type="password")
            admin_login_button = st.form_submit_button("Admin Login", type="primary")
            
            if admin_login_button:
                if admin_id and admin_password:
                    admin = verify_candidate_admin(admin_id, admin_password)
                    if admin:
                        st.session_state.user = admin
                        st.session_state.authenticated = True
                        st.session_state.user_type = 'candidate_admin'
                        st.success(f"Welcome, {admin['admin_name']}!")
                        st.rerun()
                    else:
                        st.error("Invalid admin credentials.")
                else:
                    st.error("Please fill in all fields.")



def show_assessment_window_management():
    st.title("🕒 Assessment Window Management")
    
    user = st.session_state.user
    
    # Auto-deactivate past windows
    conn = sqlite3.connect('assessment_data.db')
    cursor = conn.cursor()
    current_date = date.today()
    cursor.execute('''
        UPDATE assessment_windows 
        SET is_active = 0 
        WHERE end_date < ? AND is_active = 1
    ''', (current_date,))
    conn.commit()
    conn.close()
    
    # Create new assessment window
    st.subheader("Create New Assessment Window")
    with st.form("create_window"):
        window_name = st.text_input("Window Name")
        
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", min_value=date.today())
            start_time = st.time_input("Start Time")
        with col2:
            end_date = st.date_input("End Date",min_value=date.today())
            end_time = st.time_input("End Time")
        
        create_button = st.form_submit_button("Create Assessment Window", type="primary")
        
        if create_button:
            if window_name and start_date and end_date and start_time and end_time:
                # Additional validation for past dates and times
                current_date = date.today()
                current_time = datetime.now().time()
                
                if start_date < current_date:
                    st.error("Start date cannot be in the past.")
                elif start_date == current_date and start_time < current_time:
                    st.error("Start time cannot be in the past for today's date.")
                elif end_date < start_date:
                    st.error("End date must be after start date.")
                elif start_date == end_date and end_time <= start_time:
                    st.error("For the same date, end time must be greater than start time.")
                else:
                    if create_assessment_window(window_name, start_date, end_date, start_time, end_time, user['employee_id']):
                        st.success("Assessment window created successfully!")
                        st.rerun()
                    else:
                        st.error("Failed to create assessment window.")
            else:
                st.error("Please fill in all fields.")
    
    # Display existing windows
    st.subheader("Existing Assessment Windows")
    conn = sqlite3.connect('assessment_data.db')
    
    # Get windows with assessment counts
    windows_df = pd.read_sql_query('''
        SELECT aw.*, 
               COALESCE(COUNT(a.id), 0) as assessment_count
        FROM assessment_windows aw
        LEFT JOIN assessments a ON aw.id = a.window_id
        GROUP BY aw.id
        ORDER BY aw.created_at DESC
    ''', conn)
    conn.close()
    
    if not windows_df.empty:
        for _, window in windows_df.iterrows():
            with st.expander(f"📅 {window['window_name']} ({'Active' if window['is_active'] else 'Inactive'})"):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write(f"**Start:** {window['start_date']} at {window['start_time']}")
                    st.write(f"**End:** {window['end_date']} at {window['end_time']}")
                
                with col2:
                    st.write(f"**Created by:** {window['created_by']}")
                    st.write(f"**Assessments taken:** {window['assessment_count']}")
                
                with col3:
                    current_status = bool(window['is_active'])
                    new_status = st.toggle(
                        "Active", 
                        value=current_status, 
                        key=f"toggle_{window['id']}"
                    )
                    
                    if new_status != current_status:
                        toggle_assessment_window(window['id'], new_status)
                        st.success(f"Window {'activated' if new_status else 'deactivated'}")
                        st.rerun()
    else:
        st.info("No assessment windows created yet.")

def show_employee_dashboard():
    user = st.session_state.user
    st.markdown(f"""
    <div class="main-header">
        <h1>📊 My Assessment Dashboard</h1>
        <p>Welcome back, {user['employee_name']} ({user['employee_id']})</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Get user's assessment data
    conn = sqlite3.connect('assessment_data.db')
    df = pd.read_sql_query('''
        SELECT a.*, aw.window_name 
        FROM assessments a
        LEFT JOIN assessment_windows aw ON a.window_id = aw.id
        WHERE a.employee_id = ? 
        ORDER BY a.submit_date DESC, a.submit_time DESC
    ''', conn, params=(user['employee_id'],))
    conn.close()
    
    if df.empty:
        st.info("No assessment completed yet. Please take the assessment first.")
        return
    
    # Filter options
    st.subheader("Filter Assessments")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if 'submit_date' in df.columns:
            date_options = ["All"] + sorted(df['submit_date'].dropna().unique(), reverse=True)
            selected_date = st.selectbox("Filter by Submit Date", date_options)
        else:
            selected_date = "All"
    
    with col2:
        if 'window_name' in df.columns:
            window_options = ["All"] + list(df['window_name'].dropna().unique())
            selected_window = st.selectbox("Filter by Assessment Window", window_options)
        else:
            selected_window = "All"
    
    with col3:
        dept_options = ["All"] + list(df['department'].unique())
        selected_dept = st.selectbox("Filter by Department", dept_options)
    
    # Apply filters
    filtered_df = df.copy()
    if selected_date != "All":
        filtered_df = filtered_df[filtered_df['submit_date'] == selected_date]
    if selected_window != "All":
        filtered_df = filtered_df[filtered_df['window_name'] == selected_window]
    if selected_dept != "All":
        filtered_df = filtered_df[filtered_df['department'] == selected_dept]
    
    if filtered_df.empty:
        st.info("No assessments found for the selected filters.")
        return
    
    # Show latest assessment results
    latest = filtered_df.iloc[0]
    interpretations = json.loads(latest['interpretation'])
    
    # Display selected assessment info
    st.subheader("📋 Selected Assessment Details")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.info(f"**Window:** {latest['window_name'] or 'N/A'}")
    with col2:
        st.info(f"**Submit Date:** {latest['submit_date']}")
    with col3:
        st.info(f"**Submit Time:** {latest['submit_time']}")
    with col4:
        st.info(f"**Total Score:** {latest['total_score']}")
    
    # Recreate scores dictionary
    scores = {
        "Accountability": latest['accountability_score'],
        "Team Collaboration": latest['teamwork_score'],
        "Result Orientation": latest['result_orientation_score'],
        "Communication Skills": latest['communication_score'],
        "Adaptability": latest['adaptability_score'],
        "Integrity": latest['integrity_score'],
        "Conflict Resolution": latest['conflict_resolution_score']
    }
    
    # Calculate total possible scores
    total_possible = {comp: 36 for comp in scores.keys()}
    
    # Show results
    overall_assessment = "High Performer" if latest['total_score'] > 200 else "Average Performer"
    show_results(scores, interpretations, overall_assessment, total_possible)
    
    # Assessment history
    if len(filtered_df) > 1:
        st.subheader("📈 Assessment History")
        history_df = filtered_df[['window_name', 'submit_date', 'submit_time', 'total_score', 'department']].copy()
        if 'submit_date' in history_df.columns:
            history_df['submit_date'] = pd.to_datetime(history_df['submit_date']).dt.strftime('%Y-%m-%d')
        if 'submit_time' in history_df.columns:
            history_df['submit_time'] = history_df['submit_time'].astype(str).str[:8]
        st.dataframe(history_df, use_container_width=True)

def show_assessment_page():
    user = st.session_state.user
    
    # Check if there's an active assessment window
    active_window = get_active_assessment_window()
    
    if not active_window:
        st.error("⚠️ No active assessment window available. Please contact your administrator.")
        st.info("Assessment windows control when assessments can be taken. Please wait for the next assessment period.")
        return
    
    # Check if user has already taken assessment in this window
    if has_taken_assessment_in_window(user['employee_id'], active_window['id']):
        st.error("You have already submitted an assessment for this window.")
        return
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col3:
        try:
            st.image("Logo-TEL.png", width=200)
        except:
            st.write("Logo not found - please add Logo-TEL.png to project folder")
    
    st.markdown(f"""
    <div class="main-header">
        <h1>Tuaman Engineering Limited Behavioral Competency Assessment</h1>
        <p>Assessment Window: {active_window['window_name']}</p>
        <p>Employee Assessment Portal</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Language selection
    col1, col2 = st.columns([1, 1])
    with col1:
        language = st.selectbox(
            "Select Language / भाषा चुनें",
            ["en", "hi"],
            format_func=lambda x: "English" if x == "en" else "हिंदी"
        )
    
    # Pre-fill employee information
    st.subheader("Employee Information / कर्मचारी जानकारी")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        employee_id = st.text_input("Employee ID / कर्मचारी आईडी", value=user['employee_id'], disabled=True)
    with col2:
        employee_name = st.text_input("Employee Name / कर्मचारी नाम", value=user['employee_name'], disabled=True)
    with col3:
        department = st.text_input("Department / विभाग", value=user['department'], disabled=True)
    
    # Assessment instructions
    instructions = {
        "en": """
        ## Instructions
        1. Answer all questions honestly based on your typical behavior
        2. For Likert scale questions: 1=Strongly Disagree, 5=Strongly Agree
        3. For situational questions: Choose the best response
        4. For forced-choice questions: Select the option that better describes you
        5. Complete all sections before submitting
        """,
        "hi": """
        ## निर्देश
        1. अपने सामान्य व्यवहार के आधार पर सभी प्रश्नों का ईमानदारी से उत्तर दें
        2. लिकर्ट स्केल प्रश्नों के लिए: 1=बिल्कुल असहमत, 5=पूर्णतः सहमत
        3. स्थितिजन्य प्रश्नों के लिए: सबसे अच्छा उत्तर चुनें
        4. मजबूर विकल्प प्रश्नों के लिए: वह विकल्प चुनें जो आपका बेहतर वर्णन करता है
        5. जमा करने से पहले सभी अनुभाग पूरे करें
        """
    }
    
    st.markdown(instructions[language])
    
    # Assessment form
    responses = {}
    
    for competency in QUESTIONS.keys():
        st.markdown(f"""
        <div class="competency-section">
            <h3>📊 {competency}</h3>
        </div>
        """, unsafe_allow_html=True)
        
        for i, question in enumerate(QUESTIONS[competency][language]):
            st.markdown(f"""
            <div class="question-card">
                <p><strong>Q{i+1}:</strong> {question['question']}</p>
                <small>Points: {question['marks']}</small>
            </div>
            """, unsafe_allow_html=True)
            
            key = f"{competency}_{i}"
            
            if question["type"] == "likert":
                responses[key] = st.slider(
                    "Response",
                    1, 5,
                    value=None,
                    key=key,
                    help="1=Strongly Disagree, 2=Disagree, 3=Neutral, 4=Agree, 5=Strongly Agree"
                )
                st.markdown("""
                <div style="font-size: 14px; color: #333; margin-top: -10px; margin-bottom: 15px; 
                           background-color: #f0f2f6; padding: 8px; border-radius: 5px; border-left: 4px solid #ff4b4b;">
                    <strong>😠 1 - Strongly Disagree   🙁 2 - Disagree   😐 3 - Neutral   🙂 4 - Agree   😄 5 - Strongly Agree</strong>
                </div>
                """, unsafe_allow_html=True)
            
            elif question["type"] == "situational":
                responses[key] = st.radio(
                    "Choose the best response:",
                    range(len(question["options"])),
                    format_func=lambda x: question["options"][x],
                    key=key,
                    index=None
                )
            
            elif question["type"] == "forced_choice":
                responses[key] = st.radio(
                    "Choose what better describes you:",
                    [0, 1],
                    format_func=lambda x: question["options"][x],
                    key=key,
                    index=None
                )
            
            st.markdown("---")
    
    # Submit assessment
    if st.button("Submit Assessment / मूल्यांकन जमा करें", type="primary"):
        # Check again if user has already taken assessment in this window
        if has_taken_assessment_in_window(user['employee_id'], active_window['id']):
            st.error("You have already submitted an assessment for this window.")
            return
            
        if len(responses) < sum(len(QUESTIONS[comp][language]) for comp in QUESTIONS):
            st.error("Please answer all questions before submitting.")
            return
            
        # Calculate scores
        scores, total_possible = calculate_scores(responses, language)
        interpretations, overall_assessment = get_interpretation(scores, total_possible)
            
        # Get current date and time
        current_date = date.today()
        current_time = datetime.now().time().strftime('%H:%M:%S')
            
        # Save to database
        conn = sqlite3.connect('assessment_data.db')
        cursor = conn.cursor()
            
        cursor.execute('''
        INSERT INTO assessments (
            employee_id, employee_name, department, language, window_id,
            submit_date, submit_time,
            accountability_score, teamwork_score, result_orientation_score,
            communication_score, adaptability_score, integrity_score,
            conflict_resolution_score, total_score, responses, interpretation
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            employee_id, employee_name, department, language, active_window['id'],
            current_date, current_time,
            scores["Accountability"], scores["Team Collaboration"], scores["Result Orientation"],
            scores["Communication Skills"], scores["Adaptability"], scores["Integrity"],
            scores["Conflict Resolution"], sum(scores.values()),
            json.dumps(responses), json.dumps(interpretations)
        ))
            
        conn.commit()
        conn.close()
            
        # Display results
        st.success("Assessment completed successfully!")
        show_results(scores, interpretations, overall_assessment, total_possible)
        
        # AUTOMATICALLY SEND EMAIL TO ADMIN/HR
        st.info("📧 Sending assessment report to HR team...")
        
        user_data = {
            'employee_id': employee_id,
            'employee_name': employee_name,
            'department': department,
            'submit_date': str(current_date),
            'total_score': sum(scores.values())
        }
        
        # Generate PDF
        pdf_path = generate_assessment_pdf(user_data, scores, interpretations, overall_assessment, total_possible, "employee")
        
        if pdf_path and os.path.exists(pdf_path):
            pdf_name = f"Assessment_Report_{employee_id}_{current_date}.pdf"
            
            # Send email automatically
            subject = f"New Employee Assessment Submitted - {employee_name}"
            body = f"""Dear HR Team,

A new employee assessment has been submitted:

Employee ID: {employee_id}
Employee Name: {employee_name}
Department: {department}
Assessment Date: {current_date}
Assessment Time: {current_time}
Total Score: {sum(scores.values())}

Please find the detailed assessment report attached.

Best regards,
Assessment System"""
            
            success = send_email_with_attachment(subject, body, pdf_path, pdf_name)
            
            if success:
                st.success("✅ Assessment report automatically sent to HR team!")
            else:
                st.warning("⚠️ Assessment saved but email notification failed. HR team will be notified separately.")
        else:
            st.warning("⚠️ Assessment saved but PDF generation failed. HR team will be notified separately.")

def show_candidate_assessment_page():
    """Assessment page for candidates"""
    user = st.session_state.user
    
    # Check if candidate has already taken assessment
    if has_candidate_taken_assessment(user['candidate_code']):
        st.warning("⚠️ You have already completed the assessment.")
        st.info("Only one assessment per candidate is allowed.")
        return
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col3:
        try:
            st.image("Logo-TEL.png", width=200)
        except:
            st.write("Logo not found - please add Logo-TEL.png to project folder")
    
    st.markdown(f"""
    <div class="main-header">
        <h1>Tuaman Engineering Limited Behavioral Competency Assessment</h1>
        <p>Candidate Assessment Portal</p>
        <p>Position: {user['position_applied']}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Language selection
    col1, col2 = st.columns([1, 1])
    with col1:
        language = st.selectbox(
            "Select Language / भाषा चुनें",
            ["en", "hi"],
            format_func=lambda x: "English" if x == "en" else "हिंदी"
        )
    
    # Pre-fill candidate information
    st.subheader("Candidate Information / उम्मीदवार जानकारी")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        candidate_code = st.text_input("Candidate Code / उम्मीदवार कोड", value=user['candidate_code'], disabled=True)
    with col2:
        full_name = st.text_input("Full Name / पूरा नाम", value=user['full_name'], disabled=True)
    with col3:
        position_applied = st.text_input("Position Applied / आवेदित पद", value=user['position_applied'], disabled=True)
    
    # Assessment instructions
    instructions = {
        "en": """
        ## Instructions
        1. Answer all questions honestly based on your typical behavior
        2. For Likert scale questions: 1=Strongly Disagree, 5=Strongly Agree
        3. For situational questions: Choose the best response
        4. For forced-choice questions: Select the option that better describes you
        5. Complete all sections before submitting
        """,
        "hi": """
        ## निर्देश
        1. अपने सामान्य व्यवहार के आधार पर सभी प्रश्नों का ईमानदारी से उत्तर दें
        2. लिकर्ट स्केल प्रश्नों के लिए: 1=बिल्कुल असहमत, 5=पूर्णतः सहमत
        3. स्थितिजन्य प्रश्नों के लिए: सबसे अच्छा उत्तर चुनें
        4. मजबूर विकल्प प्रश्नों के लिए: वह विकल्प चुनें जो आपका बेहतर वर्णन करता है
        5. जमा करने से पहले सभी अनुभाग पूरे करें
        """
    }
    
    st.markdown(instructions[language])
    
    # Assessment form
    responses = {}
    
    for competency in QUESTIONS.keys():
        st.markdown(f"""
        <div class="competency-section">
            <h3>📊 {competency}</h3>
        </div>
        """, unsafe_allow_html=True)
        
        for i, question in enumerate(QUESTIONS[competency][language]):
            st.markdown(f"""
            <div class="question-card">
                <p><strong>Q{i+1}:</strong> {question['question']}</p>
                <small>Points: {question['marks']}</small>
            </div>
            """, unsafe_allow_html=True)
            
            key = f"{competency}_{i}"
            
            if question["type"] == "likert":
                responses[key] = st.slider(
                    "Response",
                    1, 5,
                    value=None,
                    key=key,
                    help="1=Strongly Disagree, 2=Disagree, 3=Neutral, 4=Agree, 5=Strongly Agree"
                )
                st.markdown("""
                <div style="font-size: 14px; color: #333; margin-top: -10px; margin-bottom: 15px; 
                           background-color: #f0f2f6; padding: 8px; border-radius: 5px; border-left: 4px solid #ff4b4b;">
                    <strong>😠 1 - Strongly Disagree   🙁 2 - Disagree   😐 3 - Neutral   🙂 4 - Agree   😄 5 - Strongly Agree</strong>
                </div>
                """, unsafe_allow_html=True)
            
            elif question["type"] == "situational":
                responses[key] = st.radio(
                    "Choose the best response:",
                    range(len(question["options"])),
                    format_func=lambda x: question["options"][x],
                    key=key,
                    index=None
                )
            
            elif question["type"] == "forced_choice":
                responses[key] = st.radio(
                    "Choose what better describes you:",
                    [0, 1],
                    format_func=lambda x: question["options"][x],
                    key=key,
                    index=None
                )
            
            st.markdown("---")
    
    # Submit assessment
    if st.button("Submit Assessment / मूल्यांकन जमा करें", type="primary"):
        # Check if candidate has already taken assessment
        if has_candidate_taken_assessment(user['candidate_code']):
            st.error("You have already submitted an assessment.")
            return
        
        if len(responses) < sum(len(QUESTIONS[comp][language]) for comp in QUESTIONS):
            st.error("Please answer all questions before submitting.")
            return
            
        # Calculate scores
        scores, total_possible = calculate_scores(responses, language)
        interpretations, overall_assessment = get_interpretation(scores, total_possible)
            
        # Get current date and time
        current_date = date.today()
        current_time = datetime.now().time().strftime('%H:%M:%S')
            
        # Save to database
        conn = sqlite3.connect('assessment_data.db')
        cursor = conn.cursor()
            
        cursor.execute('''
        INSERT INTO candidate_assessments (
            candidate_code, full_name, position_applied, language,
            submit_date, submit_time,
            accountability_score, teamwork_score, result_orientation_score,
            communication_score, adaptability_score, integrity_score,
            conflict_resolution_score, total_score, responses, interpretation
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            user['candidate_code'], user['full_name'], user['position_applied'], language,
            current_date, current_time,
            scores["Accountability"], scores["Team Collaboration"], scores["Result Orientation"],
            scores["Communication Skills"], scores["Adaptability"], scores["Integrity"],
            scores["Conflict Resolution"], sum(scores.values()),
            json.dumps(responses), json.dumps(interpretations)
        ))
            
        conn.commit()
        conn.close()
            
        # Display results
        st.success("Assessment completed successfully!")
        show_results(scores, interpretations, overall_assessment, total_possible)
        
        # AUTOMATICALLY SEND EMAIL TO HR TEAM
        st.info("📧 Sending assessment report to HR team...")
        
        user_data = {
            'candidate_code': user['candidate_code'],
            'full_name': user['full_name'],
            'position_applied': user['position_applied'],
            'submit_date': str(current_date),
            'total_score': sum(scores.values())
        }
        
        # Generate PDF
        pdf_path = generate_assessment_pdf(user_data, scores, interpretations, overall_assessment, total_possible, "candidate")
        
        if pdf_path and os.path.exists(pdf_path):
            pdf_name = f"Candidate_Assessment_Report_{user['candidate_code']}_{current_date}.pdf"
            
            # Send email automatically
            subject = f"New Candidate Assessment Submitted - {user['full_name']}"
            body = f"""Dear HR Team,

A new candidate assessment has been submitted:

Candidate Code: {user['candidate_code']}
Full Name: {user['full_name']}
Position Applied: {user['position_applied']}
Assessment Date: {current_date}
Assessment Time: {current_time}
Total Score: {sum(scores.values())}

Please find the detailed assessment report attached.

Best regards,
Assessment System"""
            
            success = send_email_with_attachment(subject, body, pdf_path, pdf_name)
            
            if success:
                st.success("✅ Assessment report automatically sent to HR team!")
            else:
                st.warning("⚠️ Assessment saved but email notification failed. HR team will be notified separately.")
        else:
            st.warning("⚠️ Assessment saved but PDF generation failed. HR team will be notified separately.")
def show_candidate_dashboard():
    """Dashboard for candidates to view their results"""
    user = st.session_state.user
    st.markdown(f"""
    <div class="main-header">
        <h1>📊 My Assessment Results</h1>
        <p>Welcome, {user['full_name']} ({user['candidate_code']})</p>
        <p>Position Applied: {user['position_applied']}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Get candidate's assessment data
    conn = sqlite3.connect('assessment_data.db')
    df = pd.read_sql_query('''
        SELECT * FROM candidate_assessments 
        WHERE candidate_code = ? 
        ORDER BY submit_date DESC, submit_time DESC
    ''', conn, params=(user['candidate_code'],))
    conn.close()
    
    if df.empty:
        st.info("No assessment completed yet. Please take the assessment first.")
        return
    
    # Show latest assessment results
    latest = df.iloc[0]
    interpretations = json.loads(latest['interpretation'])
    
    # Display assessment info
    st.subheader("📋 Assessment Details")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Submit Date:** {latest['submit_date']}")
    with col2:
        st.info(f"**Submit Time:** {latest['submit_time']}")
    with col3:
        st.info(f"**Total Score:** {latest['total_score']}")
    
    # Recreate scores dictionary
    scores = {
        "Accountability": latest['accountability_score'],
        "Team Collaboration": latest['teamwork_score'],
        "Result Orientation": latest['result_orientation_score'],
        "Communication Skills": latest['communication_score'],
        "Adaptability": latest['adaptability_score'],
        "Integrity": latest['integrity_score'],
        "Conflict Resolution": latest['conflict_resolution_score']
    }
    
    # Calculate total possible scores
    total_possible = {comp: 36 for comp in scores.keys()}
    
    # Show results
    overall_assessment = "High Performer" if latest['total_score'] > 200 else "Average Performer"
    show_results(scores, interpretations, overall_assessment, total_possible)

def show_candidate_admin_dashboard():
    """Dashboard for candidate admin to manage candidates and view results"""
    st.title("👥 Candidate Administration Dashboard")
    
    tab1, tab2, tab3 = st.tabs(["Candidate Management", "Assessment Results", "Analytics"])
    
    with tab1:
        st.subheader("Registered Candidates")
        
        # Load candidates data
        conn = sqlite3.connect('assessment_data.db')
        candidates_df = pd.read_sql_query('''
            SELECT c.*, 
                   CASE WHEN ca.candidate_code IS NOT NULL THEN 'Completed' ELSE 'Pending' END as assessment_status
            FROM candidates c
            LEFT JOIN candidate_assessments ca ON c.candidate_code = ca.candidate_code
            ORDER BY c.created_at DESC
        ''', conn)
        conn.close()
        
        if not candidates_df.empty:
            # Format expiry dates
            candidates_df['expires_at'] = pd.to_datetime(candidates_df['expires_at']).dt.strftime('%Y-%m-%d %H:%M')
            candidates_df['created_at'] = pd.to_datetime(candidates_df['created_at']).dt.strftime('%Y-%m-%d %H:%M')
            
            # Display candidates
            display_cols = ['candidate_code', 'full_name', 'position_applied', 'assessment_status', 'expires_at', 'is_active']
            st.dataframe(candidates_df[display_cols], use_container_width=True)
            
            # Bulk actions
            st.subheader("Bulk Actions")
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("Deactivate Expired Candidates"):
                    conn = sqlite3.connect('assessment_data.db')
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE candidates 
                        SET is_active = 0 
                        WHERE expires_at < datetime('now')
                    ''')
                    conn.commit()
                    conn.close()
                    st.success("Expired candidates deactivated!")
                    st.rerun()
            
            with col2:
                if st.button("Export Candidates List"):
                    csv = candidates_df.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
        else:
            st.info("No candidates registered yet.")
    
    with tab2:
        st.subheader("Assessment Results")
        
        # Load candidate assessment data
        conn = sqlite3.connect('assessment_data.db')
        results_df = pd.read_sql_query('''
            SELECT * FROM candidate_assessments 
            ORDER BY submit_date DESC, submit_time DESC
        ''', conn)
        conn.close()
        
        if not results_df.empty:
            # Filters
            col1, col2 = st.columns(2)
            with col1:
                position_filter = st.selectbox("Position", ["All"] + list(results_df['position_applied'].unique()))
            with col2:
                date_filter = st.date_input("Filter by Date", value=None, key="candidate_date_filter")
            
            # Apply filters
            filtered_df = results_df.copy()
            if position_filter != "All":
                filtered_df = filtered_df[filtered_df['position_applied'] == position_filter]
            if date_filter:
                filtered_df = filtered_df[pd.to_datetime(filtered_df['submit_date']).dt.date == date_filter]
            
            # Add percentage calculation
            filtered_df['percentage'] = (filtered_df['total_score'] / 252) * 100  # 252 is max possible score (36*7)
            
            # Display results
            display_cols = ['candidate_code', 'full_name', 'position_applied', 'submit_date', 'submit_time', 'total_score', 'percentage']
            display_df = filtered_df[display_cols].copy()
            display_df['percentage'] = display_df['percentage'].round(1)
            st.dataframe(display_df, use_container_width=True)
            
            # Export and Email options
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📥 Download Excel"):
                    # Create Excel file in memory with conditional formatting
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        display_df.to_excel(writer, sheet_name='Candidate_Results', index=False)
                        
                        # Get the workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Candidate_Results']
                        
                        # Create a red fill for poor performers
                        from openpyxl.styles import PatternFill
                        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        
                        # Apply red background to rows where percentage < 60
                        for row_num, (_, row_data) in enumerate(display_df.iterrows(), start=2):
                            if row_data['percentage'] < 60:
                                for col_num in range(1, len(display_df.columns) + 1):
                                    worksheet.cell(row=row_num, column=col_num).fill = red_fill
                    
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="Download Excel File",
                        data=excel_data,
                        file_name=f"candidate_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            with col2:
                if st.button("📧 Email Excel Report NOW", key="candidate_email_excel_direct", type="primary"):
                    st.info("Creating Excel report...")
                    
                    # Create Excel file in memory
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        display_df.to_excel(writer, sheet_name='Candidate_Results', index=False)
                        
                        # Get the workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Candidate_Results']
                        
                        # Create a red fill for poor performers
                        from openpyxl.styles import PatternFill
                        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        
                        # Apply red background to rows where percentage < 60
                        for row_num, (_, row_data) in enumerate(display_df.iterrows(), start=2):
                            if row_data['percentage'] < 60:
                                for col_num in range(1, len(display_df.columns) + 1):
                                    worksheet.cell(row=row_num, column=col_num).fill = red_fill
                    
                    excel_data = output.getvalue()
                    
                    # Save Excel to temporary file
                    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    temp_excel.write(excel_data)
                    temp_excel.close()
                    
                    excel_name = f"candidate_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    # Send email directly
                    subject = "Candidate Assessment Results - Excel Report"
                    body = """Dear Team,

Please find attached the complete candidate assessment results in Excel format.

This report contains all candidate assessment data with conditional formatting for easy analysis.
Records with performance below 60% are highlighted in red.

Best regards,
Assessment System"""
                    
                    st.info("Sending email...")
                    success = send_email_with_attachment(subject, body, temp_excel.name, excel_name)
                    
                    if success:
                        st.success("✅ Email sent successfully!")
                    else:
                        st.error("❌ Failed to send email!")
                    
                    # Clean up temp file
                    try:
                        os.unlink(temp_excel.name)
                    except:
                        pass
        else:
            st.info("No assessment results available yet.")
    
    with tab3:
        st.subheader("Candidate Analytics Dashboard")
        
        # Load candidate assessment data
        conn = sqlite3.connect('assessment_data.db')
        df = pd.read_sql_query('''
            SELECT * FROM candidate_assessments
            ORDER BY submit_date DESC, submit_time DESC
        ''', conn)
        conn.close()
        
        if df.empty:
            st.info("No candidate assessment data available yet.")
            return
        
        # Candidate filter
        st.subheader("Select Candidate")
        col1, col2 = st.columns(2)
        
        with col1:
            selected_candidate_code = st.selectbox(
                "Candidate Code", 
                [""] + list(df['candidate_code'].unique()),
                format_func=lambda x: "Select Candidate Code" if x == "" else x,
                key="analytics_candidate_code"
            )
        
        with col2:
            if selected_candidate_code:
                candidate_names = df[df['candidate_code'] == selected_candidate_code]['full_name'].unique()
                selected_candidate_name = st.selectbox("Candidate Name", candidate_names, key="analytics_candidate_name")
            else:
                selected_candidate_name = st.selectbox("Candidate Name", ["Select Candidate Code first"], key="analytics_candidate_name_empty")
        
        if not selected_candidate_code:
            st.info("Please select a candidate to view their assessment results.")
            return
        
        # Get candidate data
        candidate_data = df[(df['candidate_code'] == selected_candidate_code) & 
                          (df['full_name'] == selected_candidate_name)].iloc[-1]  # Latest assessment
        
        # Parse scores and interpretations
        scores = {
            'Accountability': candidate_data['accountability_score'],
            'Teamwork': candidate_data['teamwork_score'],
            'Result Orientation': candidate_data['result_orientation_score'],
            'Communication': candidate_data['communication_score'],
            'Adaptability': candidate_data['adaptability_score'],
            'Integrity': candidate_data['integrity_score'],
            'Conflict Resolution': candidate_data['conflict_resolution_score']
        }
        
        interpretations = json.loads(candidate_data['interpretation'])
        
        # Calculate total possible scores
        total_possible = {comp: 36 for comp in scores.keys()}
        
        # Display candidate info
        st.subheader(f"Assessment Results for {selected_candidate_name}")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.info(f"**Candidate Code:** {selected_candidate_code}")
        with col2:
            st.info(f"**Position Applied:** {candidate_data['position_applied']}")
        with col3:
            submit_date = candidate_data.get('submit_date', 'N/A')
            st.info(f"**Assessment Date:** {submit_date}")
        
        # Show the same results visualization as employee dashboard
        overall_assessment = "High Performer" if candidate_data['total_score'] > 200 else "Average Performer"
        show_results(scores, interpretations, overall_assessment, total_possible)

def show_results(scores, interpretations, overall_assessment, total_possible):
    st.subheader("📊 Assessment Results")
    
    # Overall score
    total_score = sum(scores.values())
    max_total = sum(total_possible.values())
    overall_percentage = (total_score / max_total) * 100
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Score", f"{total_score:.1f}/{max_total}", f"{overall_percentage:.1f}%")
    with col2:
        st.metric("Overall Assessment", overall_assessment)
    with col3:
        st.metric("Competencies Evaluated", len(scores))
    
    # Individual competency scores
    st.subheader("Competency Breakdown")
    
    # Create visualization
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=("Competency Scores", "Performance Levels", "Score Distribution", "Radar Chart"),
        specs=[[{"type": "bar"}, {"type": "pie"}],
               [{"type": "histogram"}, {"type": "scatterpolar"}]]
    )
    
    # Bar chart
    competencies = list(scores.keys())
    score_values = list(scores.values())
    percentages = [(scores[comp] / total_possible[comp]) * 100 for comp in competencies]
    
    fig.add_trace(
        go.Bar(x=competencies, y=percentages, name="Percentage Scores"),
        row=1, col=1
    )
    
    # Pie chart for performance levels
    levels = [interpretations[comp]["level"] for comp in competencies]
    level_counts = {level: levels.count(level) for level in set(levels)}
    
    fig.add_trace(
        go.Pie(labels=list(level_counts.keys()), values=list(level_counts.values())),
        row=1, col=2
    )
    
    # Histogram
    fig.add_trace(
        go.Scatter(
            x=competencies,
            y=percentages,
            mode='lines+markers',
            fill='tonexty',
            line=dict(width=3, color='rgba(55, 128, 191, 0.7)'),
            marker=dict(size=8, color='rgba(55, 128, 191, 1)'),
            name='Performance Ribbon'
        ),
        row=2, col=1
    )
    
    # Add a baseline at 0 for the ribbon effect
    fig.add_trace(
        go.Scatter(
            x=competencies,
            y=[0] * len(competencies),
            mode='lines',
            line=dict(width=0),
            showlegend=False,
            hoverinfo='skip'
        ),
        row=2, col=1
    )
    
    # Radar chart
    fig.add_trace(
        go.Scatterpolar(
            r=percentages,
            theta=competencies,
            fill='toself',
            name='Performance Profile',
            fillcolor='rgba(255, 165 , 0, 0.3)',
            line=dict(color='rgba(255, 165, 0, 1)')
        ),
        row=2, col=2
    )
    
    fig.update_layout(height=800, showlegend=False)
    st.plotly_chart(fig, use_container_width=True)
    
    # Detailed breakdown
    for competency in competencies:
        col1, col2, col3 = st.columns([2, 1, 2])
        
        with col1:
            st.markdown(f"**{competency}**")
            st.progress(interpretations[competency]["percentage"] / 100)
        
        with col2:
            st.markdown(f"""
            <div class="score-card">
                <h4>{interpretations[competency]['level']}</h4>
                <p>{interpretations[competency]['percentage']:.1f}%</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.write(interpretations[competency]["description"])

def show_dashboard_page():
    st.title("📈 Employee Dashboard")
    
    # Load data
    conn = sqlite3.connect('assessment_data.db')
    df = pd.read_sql_query('''
        SELECT a.*, aw.window_name 
        FROM assessments a
        LEFT JOIN assessment_windows aw ON a.window_id = aw.id
        ORDER BY COALESCE(a.submit_date, a.assessment_date) DESC, 
                 COALESCE(a.submit_time, '00:00:00') DESC
    ''', conn)
    conn.close()
    
    if df.empty:
        st.info("No assessment data available yet.")
        return
    
    # Employee filter
    st.subheader("Select Employee")
    col1, col2 = st.columns(2)
    
    with col1:
        selected_employee_id = st.selectbox(
            "Employee ID", 
            [""] + list(df['employee_id'].unique()),
            format_func=lambda x: "Select Employee ID" if x == "" else x
        )
    
    with col2:
        if selected_employee_id:
            employee_names = df[df['employee_id'] == selected_employee_id]['employee_name'].unique()
            selected_employee_name = st.selectbox("Employee Name", employee_names)
        else:
            selected_employee_name = st.selectbox("Employee Name", ["Select Employee ID first"])
    
    if not selected_employee_id:
        st.info("Please select an employee to view their assessment results.")
        return
    
    # Get employee data
    employee_data = df[(df['employee_id'] == selected_employee_id) & 
                      (df['employee_name'] == selected_employee_name)].iloc[-1]  # Latest assessment
    
    # Parse interpretation data
    interpretations = json.loads(employee_data['interpretation'])
    
    # Recreate scores dictionary
    scores = {
        "Accountability": employee_data['accountability_score'],
        "Team Collaboration": employee_data['teamwork_score'],
        "Result Orientation": employee_data['result_orientation_score'],
        "Communication Skills": employee_data['communication_score'],
        "Adaptability": employee_data['adaptability_score'],
        "Integrity": employee_data['integrity_score'],
        "Conflict Resolution": employee_data['conflict_resolution_score']
    }
    
    # Calculate total possible scores (assuming max 36 per competency based on question structure)
    total_possible = {comp: 36 for comp in scores.keys()}
    
    # Display employee info
    st.subheader(f"Assessment Results for {selected_employee_name}")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Employee ID:** {selected_employee_id}")
    with col2:
        st.info(f"**Department:** {employee_data['department']}")
    with col3:
        submit_date = employee_data.get('submit_date', 'N/A')
        if pd.isna(submit_date):
            submit_date = pd.to_datetime(employee_data['assessment_date']).strftime('%Y-%m-%d')
        st.info(f"**Assessment Date:** {submit_date}")
    
    # Show the same results as in submit assessment
    overall_assessment = "High Performer" if employee_data['total_score'] > 200 else "Average Performer"
    show_results(scores, interpretations, overall_assessment, total_possible)

def show_records_page():
    st.title("👥 Employee Records")
    
    # Load data with window information
    conn = sqlite3.connect('assessment_data.db')
    df = pd.read_sql_query('''
        SELECT a.*, aw.window_name 
        FROM assessments a
        LEFT JOIN assessment_windows aw ON a.window_id = aw.id
        ORDER BY a.submit_date DESC, a.submit_time DESC
    ''', conn)
    conn.close()
    
    if df.empty:
        st.info("No records available yet.")
        return
    
    # Filters
    st.subheader("🔍 Filter Records")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        selected_department = st.selectbox("Department", ["All"] + list(df['department'].unique()))
    with col2:
        selected_window = st.selectbox("Assessment Window", ["All"] + list(df['window_name'].dropna().unique()))
    with col3:
        date_filter = st.date_input("Filter by Date", value=None)
    
    # Apply filters
    filtered_df = df.copy()
    if selected_department != "All":
        filtered_df = filtered_df[filtered_df['department'] == selected_department]
    if selected_window != "All":
        filtered_df = filtered_df[filtered_df['window_name'] == selected_window]
    if date_filter:
        filtered_df = filtered_df[pd.to_datetime(filtered_df['submit_date']).dt.date == date_filter]
    
    # Display records
    st.subheader(f"Records Found: {len(filtered_df)}")
    
    # Format the dataframe for display
    display_df = filtered_df[['employee_id', 'employee_name', 'department', 'window_name',
                             'submit_date', 'submit_time', 'total_score', 'accountability_score', 
                             'teamwork_score', 'result_orientation_score', 'communication_score', 
                             'adaptability_score', 'integrity_score', 'conflict_resolution_score']].copy()
    
    # Calculate percentage
    max_total_score = 252  # 7 competencies × 36 marks each
    display_df['percentage'] = (display_df['total_score'] / max_total_score * 100).round(1)

    if not display_df.empty:
        st.subheader("📊 Performance Summary")
        col1, col2, col3, col4 = st.columns(4)
        
        total_employees = len(display_df)
        low_performers = len(display_df[display_df['percentage'] < 60])
        avg_score = display_df['total_score'].mean()
        avg_percentage = display_df['percentage'].mean()
        
        with col1:
            st.metric("Total Employees", total_employees)
        with col2:
            st.metric("Low Performers (<60%)", low_performers, f"{(low_performers/total_employees*100):.1f}%")
        with col3:
            st.metric("Average Score", f"{avg_score:.1f}")
        with col4:
            st.metric("Average Percentage", f"{avg_percentage:.1f}%")

    # Format dates and times
    if 'submit_date' in display_df.columns:
        display_df['submit_date'] = pd.to_datetime(display_df['submit_date'], errors='coerce').dt.strftime('%Y-%m-%d')
    if 'submit_time' in display_df.columns:
        display_df['submit_time'] = display_df['submit_time'].astype(str).str[:8]
    
    # Style the dataframe to highlight low performers (< 60%) in RED
    def highlight_low_performers(row):
        if row['percentage'] < 60:
            return ['background-color: #ffcccc; color: #cc0000;'] * len(row)
        else:
            return [''] * len(row)
    
    styled_df = display_df.style.apply(highlight_low_performers, axis=1)
    
    # Display the styled dataframe
    st.dataframe(styled_df, use_container_width=True)
    
    # Add legend
    st.markdown("""
    <div style="background-color: #f0f2f6; padding: 10px; border-radius: 5px; margin: 10px 0;">
        <strong>Legend:</strong> 
        <span style="background-color: #ffcccc; color: #cc0000; padding: 2px 8px; border-radius: 3px; font-weight: bold;">
            Red rows indicate performance below 60%
        </span>
    </div>
    """, unsafe_allow_html=True)
    
    # Export and Email functionality
    st.subheader("📊 Export & Email Options")
    
    col1, col2 = st.columns([1,1])
    
    with col1:
        if st.button("📥 Download Excel"):
            # Create Excel file in memory with conditional formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_df.to_excel(writer, sheet_name='Assessment_Records', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Assessment_Records']
                
                # Create a red fill for poor performers
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                # Apply red background to rows where percentage < 60
                for row_num, (_, row_data) in enumerate(display_df.iterrows(), start=2):
                    if row_data['percentage'] < 60:
                        for col_num in range(1, len(display_df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = red_fill
            
            excel_data = output.getvalue()
            
            st.download_button(
                label="Download Excel File",
                data=excel_data,
                file_name=f"assessment_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    with col2:
        if st.button("📧 Email Excel Report NOW", key="admin_email_excel_direct", type="primary"):
            st.info("Creating Excel report...")
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_df.to_excel(writer, sheet_name='Assessment_Records', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Assessment_Records']
                
                # Create a red fill for poor performers
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                # Apply red background to rows where percentage < 60
                for row_num, (_, row_data) in enumerate(display_df.iterrows(), start=2):
                    if row_data['percentage'] < 60:
                        for col_num in range(1, len(display_df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = red_fill
            
            excel_data = output.getvalue()
            
            # Save Excel to temporary file
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_excel.write(excel_data)
            temp_excel.close()
            
            excel_name = f"assessment_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Send email directly
            subject = "Employee Assessment Records - Excel Report"
            body = """Dear Team,

Please find attached the complete employee assessment records in Excel format.

This report contains all assessment data with conditional formatting for easy analysis.
Records with performance below 60% are highlighted in red.

Best regards,
Assessment System"""
            
            st.info("Sending email...")
            success = send_email_with_attachment(subject, body, temp_excel.name, excel_name)
            
            if success:
                st.success("✅ Email sent successfully!")
            else:
                st.error("❌ Failed to send email!")
            
            # Clean up temp file
            try:
                os.unlink(temp_excel.name)
            except:
                pass
        
    
    # DIRECT EMAIL SENDING - NO PREVIEW
    
    
def main():
    # Initialize database
    st.set_page_config(
        page_title="Talent Gateway",
        page_icon="Logo-TEL.png",  # This will use your Tuaman logo as favicon

    )
    init_database()
    
    # Custom CSS
    st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 2rem 0;
        background: linear-gradient(90deg, #1f4e79, #2e86ab);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .competency-section {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    .question-card {
        background-color: white;
        padding: 1rem;
        border-left: 4px solid #1f4e79;
        margin: 0.5rem 0;
        border-radius: 5px;
    }
    .score-card {
        text-align: center;
        padding: 1rem;
        background-color: #f8f9fa;
        border-radius: 8px;
        border: 1px solid #dee2e6;
    }
    .stButton > button {
        width: 100%;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Initialize session state
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'user' not in st.session_state:
        st.session_state.user = None
    if 'user_category' not in st.session_state:
        st.session_state.user_category = None
    if 'user_type' not in st.session_state:
        st.session_state.user_type = None
    show_footer()
    
    # Show initial selection if no category chosen
    if not st.session_state.user_category:
        show_initial_selection()
        return
    
    # Show appropriate login page based on category
    if not st.session_state.authenticated:
        if st.session_state.user_category == "employee":
            show_login_page()
        elif st.session_state.user_category == "candidate":
            show_candidate_login_page()
        return
    
    # Main application logic based on user type
    user = st.session_state.user
    user_type = st.session_state.user_type
    
    # Sidebar navigation
    with st.sidebar:
        st.markdown(f"**Logged in as:** {user.get('employee_name', user.get('full_name', user.get('admin_name', 'User')))}")
        st.markdown(f"**Type:** {user_type.replace('_', ' ').title()}")
        
        if st.button("Logout"):
            st.session_state.authenticated = False
            st.session_state.user = None
            st.session_state.user_category = None
            st.session_state.user_type = None
            st.rerun()
    
    # Navigation based on user type
    if user_type == "employee":
        if user.get('user_type') == 'admin':
            # Admin navigation
            page = st.sidebar.selectbox(
                "Navigation",
                ["View Dashboard", "Employee Records", "Assessment Windows"]
            )
            
            if page == "View Dashboard":
                show_dashboard_page()
            elif page == "Employee Records":
                show_records_page()
            elif page == "Assessment Windows":
                show_assessment_window_management()
        else:
            # Regular employee navigation
            page = st.sidebar.selectbox(
                "Navigation",
                ["Take Assessment", "My Dashboard"]
            )
            
            if page == "Take Assessment":
                show_assessment_page()
            elif page == "My Dashboard":
                show_employee_dashboard()
    
    elif user_type == "candidate":
        # Candidate navigation
        page = st.sidebar.selectbox(
            "Navigation",
            ["Take Assessment", "My Results"]
        )
        
        if page == "Take Assessment":
            show_candidate_assessment_page()
        elif page == "My Results":
            show_candidate_dashboard()
    
    elif user_type == "candidate_admin":
        # Candidate admin navigation
        show_candidate_admin_dashboard()

if __name__ == "__main__":
    main()
